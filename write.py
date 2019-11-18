from openpyxl import load_workbook
from collections import OrderedDict
import argparse
import re


def set_filename(filename, add_str=''):
    match_str = re.match(r"(.*), г\. (\w+), ([а-яА-ЯёЁ\.\-\d ]+), (.*), (.*)", add_str)
    if not match_str:
        match_str = re.match(
            r"(.*), г\. (\w+), ([а-яА-ЯёЁ\.\-\d ]+), (.*)", add_str)
    add_str_filename = match_str.group(3).replace('.', '')\
        .replace(' ', '_').replace('/', '_').strip()
    home = match_str.group(4).replace('.', '_').replace(' ', '').replace('/', '_').strip()
    return '{}_{}_{}.xlsx'.format(filename.rsplit('.', 1)[0], add_str_filename, home)


def save_new_files_with_new_cell_value(
        donor_filename,
        donor_obj,
        accept_filename,
        accept_cell,
        output_dir='./',
        set_filename_f=set_filename,
):
    wb1 = load_workbook(filename=donor_filename)
    ws11 = wb1.active
    donor_dict = OrderedDict()

    for cellObj in ws11[donor_obj[0]:donor_obj[1]]:
        for cell in cellObj:
            donor_dict[cell.coordinate] = cell.value

    wb2 = load_workbook(filename=accept_filename)
    ws21 = wb2.active
    # ws21.title = "range names"
    filename_set = set()

    for coordinate, cell_value in donor_dict.items():
        ws21[accept_cell] = cell_value
        new_accept_filename = set_filename_f(accept_filename, cell_value)
        filename_set.add(new_accept_filename)
        wb2.save(filename='{}{}'.format(output_dir, new_accept_filename))

    print('сохраненных файлов должно быть: {}'.format(len(filename_set)))


def set_cli_argument_parse():
    parser = argparse.ArgumentParser(
        description="Для создания excel файлов с заменой ячейки на значения "
                    "ячеек выбраного диапазоана "
    )

    parser.add_argument("-dfile", "--donor_file", dest="donor_file",
                        help="устанавливает имя файла откуда берется диапазон")
    parser.add_argument('-dcells', '--donor_cells', nargs=2, dest="donor_cells",
                        help='Диапазон ячеек для подстановки в новые файлы')
    parser.add_argument("-afile", "--accept_file", dest="accept_file",
                        help="устанавливает имя файла куда "
                             "подставляются значения"
                        )
    parser.add_argument("-acell", "--accept_cell", dest="accept_cell",
                        help="устанавливает ячейку для замены значения")
    parser.add_argument("-odir", "--output_dir",
                        dest="output_dir",
                        default='tech_maps/',
                        help="директория для сохранения новых файлов, "
                        )

    return parser.parse_args()


if __name__ == '__main__':
    cli_argument_parser = set_cli_argument_parse()

    m = re.match(r"(.*), г\. (\w+), ([а-яА-ЯёЁ\.\-\d ]+), (.*), (.*)",
                 "127051, г. Москва, пер. Колобовский 3-й, д. 8, стр. 3")


    print(set_filename(
        'Техкарта.xlsx', "141408, г. Москва, дер-ня Новосёлки, проезд Охтинский, д. 6")
    )

    # save_new_files_with_new_cell_value(
    #     'address.xlsx', ['C2', 'C207'], 'Техкарта.xlsx', 'D5')
    # python write.py -dfile 'address.xlsx' -dcells 'C2' 'C207' -afile 'Техкарта.xlsx' -acell 'D5' -odir 'tech_maps/'

    save_new_files_with_new_cell_value(
        cli_argument_parser.donor_file,
        cli_argument_parser.donor_cells,
        cli_argument_parser.accept_file,
        cli_argument_parser.accept_cell,
        cli_argument_parser.output_dir,
    )

